from __future__ import annotations

from pathlib import Path

from app.models.participant import Participant


class ParticipantRepository:
    def load_from_excel(self, excel_path: Path) -> list[Participant]:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to load participants from Excel.") from exc

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active

        participants = []
        required_columns = {"email", "first_name", "last_name"}
        column_indices = {}

        # Parse header row
        header_row = sheet[1]
        for col_idx, cell in enumerate(header_row, 1):
            col_name = cell.value
            if col_name in required_columns:
                column_indices[col_name] = col_idx

        missing_columns = required_columns - set(column_indices.keys())
        if missing_columns:
            raise ValueError(f"Missing required columns in participant Excel: {missing_columns}")

        # Parse data rows
        seen_emails = set()
        for row_idx in range(2, sheet.max_row + 1):
            email = sheet.cell(row_idx, column_indices["email"]).value
            first_name = sheet.cell(row_idx, column_indices["first_name"]).value
            last_name = sheet.cell(row_idx, column_indices["last_name"]).value

            if not email or not first_name or not last_name:
                continue

            email_str = str(email).strip()
            if email_str in seen_emails:
                raise ValueError(f"Duplicate email in participant Excel: {email_str}")
            seen_emails.add(email_str)

            participants.append(
                Participant(
                    email=email_str,
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                )
            )

        return participants
