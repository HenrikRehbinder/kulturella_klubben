from __future__ import annotations

from pathlib import Path

from app.models.prize import Prize


class PrizeRepository:
    def load_from_excel(self, excel_path: Path) -> list[Prize]:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to load prizes from Excel.") from exc

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active

        prizes = []
        required_columns = {"prize_id", "title", "description", "image_path"}
        column_indices = {}

        # Parse header row
        header_row = sheet[1]
        for col_idx, cell in enumerate(header_row, 1):
            col_name = cell.value
            if col_name in required_columns:
                column_indices[col_name] = col_idx

        missing_columns = required_columns - set(column_indices.keys())
        if missing_columns:
            raise ValueError(f"Missing required columns in prize Excel: {missing_columns}")

        # Parse data rows
        seen_ids = set()
        for row_idx in range(2, sheet.max_row + 1):
            prize_id = sheet.cell(row_idx, column_indices["prize_id"]).value
            title = sheet.cell(row_idx, column_indices["title"]).value
            description = sheet.cell(row_idx, column_indices["description"]).value
            image_path = sheet.cell(row_idx, column_indices["image_path"]).value

            if not prize_id or not title or not description or not image_path:
                continue

            prize_id_str = str(prize_id).strip()
            if prize_id_str in seen_ids:
                raise ValueError(f"Duplicate prize_id in prize Excel: {prize_id_str}")
            seen_ids.add(prize_id_str)

            image_file = Path(str(image_path).strip())
            if not image_file.exists():
                raise ValueError(f"Prize image file not found: {image_path}")

            prizes.append(
                Prize(
                    prize_id=prize_id_str,
                    title=str(title).strip(),
                    description=str(description).strip(),
                    image_path=str(image_path).strip(),
                    is_available=True,
                )
            )

        return prizes
