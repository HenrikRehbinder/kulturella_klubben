from pathlib import Path

import pytest

from app.services.participant_repository import ParticipantRepository
from app.services.prize_repository import PrizeRepository


@pytest.fixture
def tmp_excel_with_participants(tmp_path: Path) -> Path:
    """Create a test Excel file with participants."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    excel_file = tmp_path / "participants.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["email", "first_name", "last_name"])
    sheet.append(["alice@example.com", "Alice", "Alpha"])
    sheet.append(["bob@example.com", "Bob", "Beta"])
    workbook.save(excel_file)
    return excel_file


@pytest.fixture
def tmp_excel_with_prizes(tmp_path: Path) -> Path:
    """Create a test Excel file with prizes."""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    # Create test image files
    image1 = tmp_path / "prize1.png"
    image1.write_text("fake image 1")
    image2 = tmp_path / "prize2.png"
    image2.write_text("fake image 2")

    excel_file = tmp_path / "prizes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["prize_id", "title", "description", "image_path"])
    sheet.append(["p1", "Prize One", "The first prize", str(image1)])
    sheet.append(["p2", "Prize Two", "The second prize", str(image2)])
    workbook.save(excel_file)
    return excel_file


def test_load_participants(tmp_excel_with_participants: Path) -> None:
    repo = ParticipantRepository()
    participants = repo.load_from_excel(tmp_excel_with_participants)

    assert len(participants) == 2
    assert participants[0].email == "alice@example.com"
    assert participants[0].first_name == "Alice"
    assert participants[1].email == "bob@example.com"


def test_load_participants_duplicate_email_raises_error(tmp_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    excel_file = tmp_path / "participants.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["email", "first_name", "last_name"])
    sheet.append(["alice@example.com", "Alice", "Alpha"])
    sheet.append(["alice@example.com", "Alice", "Other"])
    workbook.save(excel_file)

    repo = ParticipantRepository()
    with pytest.raises(ValueError, match="Duplicate email"):
        repo.load_from_excel(excel_file)


def test_load_participants_missing_columns_raises_error(tmp_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    excel_file = tmp_path / "participants.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["email", "first_name"])
    sheet.append(["alice@example.com", "Alice"])
    workbook.save(excel_file)

    repo = ParticipantRepository()
    with pytest.raises(ValueError, match="Missing required columns"):
        repo.load_from_excel(excel_file)


def test_load_prizes(tmp_excel_with_prizes: Path) -> None:
    repo = PrizeRepository()
    prizes = repo.load_from_excel(tmp_excel_with_prizes)

    assert len(prizes) == 2
    assert prizes[0].prize_id == "p1"
    assert prizes[0].title == "Prize One"
    assert prizes[1].prize_id == "p2"


def test_load_prizes_missing_image_raises_error(tmp_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    excel_file = tmp_path / "prizes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["prize_id", "title", "description", "image_path"])
    sheet.append(["p1", "Prize One", "Description", "/nonexistent/image.png"])
    workbook.save(excel_file)

    repo = PrizeRepository()
    with pytest.raises(ValueError, match="Prize image file not found"):
        repo.load_from_excel(excel_file)
