from pathlib import Path

import pytest

from app.services.mock_data_generator import MockDataGenerator


def test_generate_mock_participants(tmp_path: Path) -> None:
    """Test mock participant generation."""
    output_file = tmp_path / "participants.xlsx"
    MockDataGenerator.generate_mock_participants(output_file, count=10)

    assert output_file.exists()

    # Verify content
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    workbook = load_workbook(output_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert len(rows) == 11  # Header + 10 data rows
    assert rows[0] == ("email", "first_name", "last_name")
    assert "@kulturella.se" in rows[1][0]


def test_generate_mock_prize_images(tmp_path: Path) -> None:
    """Test mock prize image generation."""
    image_dir = tmp_path / "images"
    MockDataGenerator.generate_mock_prize_images(image_dir, count=5)

    assert image_dir.exists()
    assert len(list(image_dir.glob("*.png"))) == 5


def test_generate_mock_prizes(tmp_path: Path) -> None:
    """Test mock prize generation."""
    output_file = tmp_path / "prizes.xlsx"
    image_dir = tmp_path / "images"
    image_dir.mkdir()

    # Create dummy images first
    for i in range(5):
        (image_dir / f"prize_{i + 1:02d}.png").write_text("dummy")

    MockDataGenerator.generate_mock_prizes(output_file, image_dir, count=5)

    assert output_file.exists()

    # Verify content
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    workbook = load_workbook(output_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert len(rows) == 6  # Header + 5 data rows
    assert rows[0] == ("prize_id", "title", "description", "image_path")


def test_generate_all_mock_data(tmp_path: Path) -> None:
    """Test complete mock data generation."""
    MockDataGenerator.generate_all(tmp_path, session_date="2026-05-06", participant_count=5, prize_count=3)

    data_dir = tmp_path / "data"
    assert (data_dir / "participants" / "participants.xlsx").exists()
    assert (data_dir / "sessions" / "2026-05-06" / "prizes.xlsx").exists()
    assert len(list((data_dir / "sessions" / "2026-05-06" / "prize_images").glob("*.png"))) == 3
